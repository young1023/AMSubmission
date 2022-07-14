import os
import io
import msoffcrypto
import openpyxl
from openpyxl import Workbook
import pandas as pd
from datetime import datetime
import gnupg
import shutil
import openpyxl

# current date and time
now = datetime.now()

# Get current path of the script file
currentPath = os.path.dirname(os.path.realpath(__file__))

exeFolder    = currentPath+'/file/'



def Excel2Text(exeFolder,currentPath):


        # check if file exist
        if os.listdir(exeFolder):
               
                # Search the folder for Excel file
                for file in os.listdir(exeFolder):

                         file = exeFolder + file

                         wb = openpyxl.load_workbook(file)

                         print(wb.worksheets[0].protection._password)
                                    
                # Define a BytesIO object 
                decrypted_workbook = io.BytesIO()

                # Open the file with the following conditions
                with open(file, 'rb') as file:
       
                         office_file = msoffcrypto.OfficeFile(file)

                         office_file.load_key(password='Shell1234')
    
                         office_file.decrypt(decrypted_workbook)

                # Open the Excel file
                workbook = openpyxl.load_workbook(filename=decrypted_workbook)

 
                worksheet = workbook.active 

                df = pd.DataFrame(worksheet.values)

                print(df) #for debug

                df.dropna(axis=0, how='all', inplace=True)

                print(df) #for debug

                row = df.shape[0]  

                # Set target file , get current year, month, date
                filename = 'SZ4.txt'

                # Set the header of file
                header = 'HDNONAIR' + 4*' ' + now.strftime("%Y") + now.strftime("%m") + now.strftime("%d") + 'SZ4 ' + 175*' ' + '.' + '\n'

                # Define the text file 
                with open(exeFolder + filename,'w',encoding='UTF8', newline='') as f:

                        # write the header
                        f.write(header)

                 # Define the text file 
                with open(exeFolder + filename,'a',encoding='UTF8', newline='') as f:

                
                        #for row in worksheet.rows:
                        for i in range(1,df.shape[0]+1):
                            
                        
                            # Define the csv file of HKT
                            with open(exeFolder + filename,'a') as f:

                                miles = str(worksheet.cell(row=i,column=8).value)

                                

                                if str(worksheet.cell(row=i,column=12).value) != 'None':

                                        
                                        if i > 1:
                                        
                                                # write the data to file, member number (12 characters)
                                                f.write('AC' + str(worksheet.cell(row=i,column=12).value)  + \
                                                (10-int(len(str(worksheet.cell(row=i,column=12).value))))*' ' + \
                                                # Family name (25 characters)
                                                str(worksheet.cell(row=i,column=6).value)[:25]  + (25-int(len(str(worksheet.cell(row=i,column=6).value))))*' ' + \
                                                # Given name(25 characters) + 25 characters filter
                                                str(worksheet.cell(row=i,column=5).value)[:25] + (25-int(len(str(worksheet.cell(row=i,column=5).value))))*' ' + 25*' ' + \
                                                # Check-in date, (Activity Desc + Check-out date - 38 space)
                                                str(worksheet.cell(row=i,column=11).value)[6:] + str(worksheet.cell(row=i,column=11).value)[3:-5] + \
                                                str(worksheet.cell(row=i,column=11).value)[:2] + 38*' ' + \
                                                # Mile
                                                (8-int(len(miles)))*'0' + miles + \
                                                # Partnet code + 5 characters filter + 
                                                '0903270001' + 5*' ' + 'SHELLHKCON' +  \
                                                #Asia Miles reference code
                                                (15-len(str(worksheet.cell(row=i,column=3).value)))*'0' + \
                                                str(worksheet.cell(row=i,column=3).value)  + 18*' ' + '.' + '\n')

                        
                                
                        # Define the text file 
                        with open(exeFolder + filename,'a',encoding='UTF8', newline='') as f:
                        
                           

                                # write the footer
                                f.write('$$' + (6-len(str(row-1)))*'0' + str(row-1) + \

                        (6-len(str(row-1)))*'0' + str(row-1) + 24*'0' + 161*' ' + '.')

                gpg = gnupg.GPG()

                # import key
                with open(currentPath+'\key\cx_prod_key.asc') as f:
                
                    key_data = f.read()
            
                import_result = gpg.import_keys(key_data)

                # encrypt file
 
                with open(exeFolder+'SZ4.txt', 'rb') as f:

                        status = gpg.encrypt_file(
                        
                        f, recipients=['imt#mktb@cathaypacific.com'],
                
                        output=exeFolder+'SZ4.txt.gpg')


       

        
#Excel2Text(exeFolder,currentPath)

        
                

      
